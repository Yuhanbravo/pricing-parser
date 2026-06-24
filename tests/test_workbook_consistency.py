from __future__ import annotations

import csv
import math
import re
from pathlib import Path

import pytest
from openpyxl import load_workbook

from valuation_parser.exporters import (
    POSITION_FIELDS,
    REVIEW_FIELDS,
    ROUTING_FIELDS,
    SUBJECT_FIELDS,
    write_excel_workbook,
)
from valuation_parser.models import PositionRecord, ReviewItem, RouteDecision, SubjectRecord
from valuation_parser.pipeline import run_pipeline


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _cell_values_equal(workbook_val: object | None, csv_val: str) -> bool:
    """Compare a workbook cell value with a CSV cell value.

    Rules:
    - ``None`` (workbook) and ``""`` (CSV) are equivalent empty values.
    - Numeric values are compared as floats with ``math.isclose(rel_tol=1e-9)``
      to tolerate int/float representation differences.
    - All other values are compared as strings.
    """
    # --- empty-value equivalence ---
    if workbook_val is None and csv_val == "":
        return True
    if workbook_val is None and csv_val != "":
        return False
    if workbook_val is not None and csv_val == "":
        return False

    # Both sides are non-empty.
    # --- numeric comparison ---
    try:
        wv_float = float(workbook_val)  # type: ignore[arg-type]
        cv_float = float(csv_val)
        return math.isclose(wv_float, cv_float, rel_tol=1e-9)
    except (ValueError, TypeError):
        pass

    # --- string comparison ---
    return str(workbook_val) == csv_val


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    """Read a CSV file and return its rows as list[dict] (header → value)."""
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_workbook_sheets(path: Path) -> dict[str, list[list[object | None]]]:
    """Read all sheets from a workbook.

    Returns:
        ``{sheet_name: [rows]}`` where each row is a list of cell values.
        The first row of each sheet is the header row.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            sheet_name: [list(row) for row in wb[sheet_name].iter_rows(values_only=True)]
            for sheet_name in wb.sheetnames
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Expected sheet names (in order)
# ---------------------------------------------------------------------------

EXPECTED_SHEET_NAMES = [
    "routing_results",
    "valuation_subjects",
    "valuation_positions",
    "review_items",
]

SHEET_CSV_MAP = {
    "routing_results": "routing_results",
    "valuation_subjects": "valuation_subjects",
    "valuation_positions": "valuation_positions",
    "review_items": "review_items",
}

SHEET_FIELD_MAP = {
    "routing_results": ROUTING_FIELDS,
    "valuation_subjects": SUBJECT_FIELDS,
    "valuation_positions": POSITION_FIELDS,
    "review_items": REVIEW_FIELDS,
}


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module")
def pipeline_outputs(tmp_path_factory: pytest.TempPathFactory) -> dict[str, Path]:
    """Run the full pipeline once on ``data_samples/raw/`` (module-scoped).

    .. warning::
        All tests using this fixture must treat the output files as **read-only**.
        The fixture is module-scoped for performance; mutating the shared output
        directory will cause cross-test contamination. If you need to modify
        output files, use a separate ``tmp_path``-scoped fixture instead.
    """
    output_dir = tmp_path_factory.mktemp("output_consistency")
    return run_pipeline(
        Path("data_samples/raw"),
        Path("产品与托管机构映射表.csv"),
        output_dir,
    )


# ---------------------------------------------------------------------------
# Test 1 — Sheet names
# ---------------------------------------------------------------------------

def test_workbook_sheet_names_match_expected_set(pipeline_outputs: dict[str, Path]) -> None:
    """Workbook must contain exactly the four expected sheets, in order."""
    sheets = _read_workbook_sheets(pipeline_outputs["output_workbook"])
    assert list(sheets.keys()) == EXPECTED_SHEET_NAMES, (
        f"Expected sheet names {EXPECTED_SHEET_NAMES}, got {list(sheets.keys())}"
    )


# ---------------------------------------------------------------------------
# Test 2 — Headers match field constants
# ---------------------------------------------------------------------------

def test_workbook_headers_match_field_constants(pipeline_outputs: dict[str, Path]) -> None:
    """Every workbook sheet header row must match its ``*_FIELDS`` constant."""
    sheets = _read_workbook_sheets(pipeline_outputs["output_workbook"])

    for sheet_name, expected_fields in SHEET_FIELD_MAP.items():
        header_row = sheets[sheet_name][0]
        # openpyxl reads values in their native Python types; normalise to str
        header_values = [str(v) for v in header_row]
        assert header_values == expected_fields, (
            f"{sheet_name} header mismatch:\n"
            f"  expected ({len(expected_fields)} cols): {expected_fields}\n"
            f"  got      ({len(header_values)} cols): {header_values}"
        )


# ---------------------------------------------------------------------------
# Test 3 — Row counts match CSV
# ---------------------------------------------------------------------------

def test_workbook_row_counts_match_csv(pipeline_outputs: dict[str, Path]) -> None:
    """Workbook data row count must match the corresponding CSV data row count."""
    sheets = _read_workbook_sheets(pipeline_outputs["output_workbook"])

    for sheet_name, csv_name in SHEET_CSV_MAP.items():
        wb_data_rows = len(sheets[sheet_name]) - 1  # exclude header row
        csv_rows = _read_csv_rows(pipeline_outputs[csv_name])
        csv_data_rows = len(csv_rows)
        assert wb_data_rows == csv_data_rows, (
            f"{sheet_name}: workbook has {wb_data_rows} data rows "
            f"but {csv_name} has {csv_data_rows}"
        )


# ---------------------------------------------------------------------------
# Test 4 — Selected cell values match CSV
# ---------------------------------------------------------------------------

def test_workbook_first_middle_last_rows_match_csv(pipeline_outputs: dict[str, Path]) -> None:
    """First / middle / last data rows in each workbook sheet must match CSV.

    Covers:
    - deterministic row sampling (not random)
    - null/empty equivalence (``None`` ↔ ``""``)
    - numeric int/float tolerance
    """
    sheets = _read_workbook_sheets(pipeline_outputs["output_workbook"])

    for sheet_name, csv_name in SHEET_CSV_MAP.items():
        wb_rows = sheets[sheet_name]
        wb_header = [str(v) for v in wb_rows[0]]
        wb_data = wb_rows[1:]

        csv_rows = _read_csv_rows(pipeline_outputs[csv_name])

        if not wb_data:
            continue

        # Deterministic sampling: first, middle, last
        sample_indices: set[int] = {0}
        if len(wb_data) >= 2:
            sample_indices.add(len(wb_data) - 1)  # last
        if len(wb_data) >= 3:
            sample_indices.add(len(wb_data) // 2)  # middle

        for row_idx in sorted(sample_indices):
            wb_row = wb_data[row_idx]
            csv_row = csv_rows[row_idx]

            for col_idx, field_name in enumerate(wb_header):
                wb_val = wb_row[col_idx] if col_idx < len(wb_row) else None
                csv_val = csv_row.get(field_name, "")

                assert _cell_values_equal(wb_val, csv_val), (
                    f"{sheet_name}[row {row_idx + 1}, col '{field_name}']: "
                    f"workbook={wb_val!r} ≠ CSV={csv_val!r}"
                )


def test_workbook_null_empty_equivalence(pipeline_outputs: dict[str, Path]) -> None:
    """Rows with known nullable fields must treat workbook ``None`` and CSV ``""`` as equal.

    This test specifically targets rows in review_items where optional fields
    (e.g. quantity, cost, market_value, pnl) are often ``None``.
    """
    sheets = _read_workbook_sheets(pipeline_outputs["output_workbook"])

    # Focus on review_items — it has the most optional numeric fields
    wb_rows = sheets["review_items"]
    wb_header = [str(v) for v in wb_rows[0]]
    wb_data = wb_rows[1:]

    csv_rows = _read_csv_rows(pipeline_outputs["review_items"])

    # Fields that are commonly None
    nullable_fields = {"quantity", "cost", "market_value", "pnl", "review_note"}

    null_found = False
    for row_idx, wb_row in enumerate(wb_data):
        csv_row = csv_rows[row_idx]
        for col_idx, field_name in enumerate(wb_header):
            if field_name not in nullable_fields:
                continue
            wb_val = wb_row[col_idx] if col_idx < len(wb_row) else None
            csv_val = csv_row.get(field_name, "")
            if wb_val is None or csv_val == "":
                null_found = True
                assert _cell_values_equal(wb_val, csv_val), (
                    f"review_items[row {row_idx + 1}, col '{field_name}']: "
                    f"null/empty mismatch: workbook={wb_val!r} ≠ CSV={csv_val!r}"
                )

    # Sanity check: at least one null/empty value was verified
    assert null_found, (
        "Expected at least one null/empty value in review_items nullable fields, found none"
    )


def test_workbook_numeric_cells_match_csv(pipeline_outputs: dict[str, Path]) -> None:
    """Numeric cells in valuation_positions must match CSV with float tolerance.

    This test verifies that int/float representation differences between
    openpyxl and CSV serialization are handled correctly.
    """
    sheets = _read_workbook_sheets(pipeline_outputs["output_workbook"])

    wb_rows = sheets["valuation_positions"]
    wb_header = [str(v) for v in wb_rows[0]]
    wb_data = wb_rows[1:]

    csv_rows = _read_csv_rows(pipeline_outputs["valuation_positions"])

    # Fields expected to carry numeric values
    numeric_fields = {
        "quantity", "unit_cost", "cost", "market_price",
        "market_value", "unrealized_pnl",
    }

    numeric_found = False
    for row_idx, wb_row in enumerate(wb_data):
        csv_row = csv_rows[row_idx]
        for col_idx, field_name in enumerate(wb_header):
            if field_name not in numeric_fields:
                continue
            wb_val = wb_row[col_idx] if col_idx < len(wb_row) else None
            csv_val = csv_row.get(field_name, "")

            # Only check cells that actually have values on both sides
            if wb_val is not None and csv_val != "":
                numeric_found = True
                try:
                    wv_float = float(wb_val)  # type: ignore[arg-type]
                    cv_float = float(csv_val)
                    assert math.isclose(wv_float, cv_float, rel_tol=1e-9), (
                        f"valuation_positions[row {row_idx + 1}, col '{field_name}']: "
                        f"numeric mismatch: workbook={wb_val!r} ≠ CSV={csv_val!r}"
                    )
                except (ValueError, TypeError):
                    # Non-numeric value in a numeric field — fall back to string comparison
                    assert str(wb_val) == csv_val, (
                        f"valuation_positions[row {row_idx + 1}, col '{field_name}']: "
                        f"string mismatch: workbook={wb_val!r} ≠ CSV={csv_val!r}"
                    )

    assert numeric_found, (
        "Expected at least one numeric value in valuation_positions, found none"
    )


# ---------------------------------------------------------------------------
# Test 5 — Date-derived filename stability
# ---------------------------------------------------------------------------

WORKBOOK_FILENAME_PATTERN = re.compile(r"估值表解析_output_\d{4}-\d{2}-\d{2}\.xlsx")


def test_workbook_filename_matches_date_pattern(pipeline_outputs: dict[str, Path]) -> None:
    """Workbook filename must match ``估值表解析_output_<YYYY-MM-DD>.xlsx``."""
    name = pipeline_outputs["output_workbook"].name
    assert WORKBOOK_FILENAME_PATTERN.fullmatch(name), (
        f"Workbook filename '{name}' does not match expected date pattern"
    )


def test_workbook_filename_uses_earliest_input_date(pipeline_outputs: dict[str, Path]) -> None:
    """Workbook filename date must be the earliest date found across all input files.

    The full ``data_samples/raw/`` set contains files all dated 2025-03-27,
    so the expected filename is ``估值表解析_output_2025-03-27.xlsx``.
    """
    name = pipeline_outputs["output_workbook"].name
    assert name == "估值表解析_output_2025-03-27.xlsx", (
        f"Expected workbook filename '估值表解析_output_2025-03-27.xlsx', got '{name}'"
    )


def test_workbook_filename_deterministic(pipeline_outputs: dict[str, Path]) -> None:
    """Running the pipeline twice on the same inputs must produce the same filename."""
    # Derive a second output dir from the first run's temp tree
    first_output_dir = pipeline_outputs["output_workbook"].parent
    second_output_dir = first_output_dir.parent / "output_run2"

    outputs2 = run_pipeline(
        Path("data_samples/raw"),
        Path("产品与托管机构映射表.csv"),
        second_output_dir,
    )
    assert outputs2["output_workbook"].name == pipeline_outputs["output_workbook"].name, (
        f"Filename changed across runs: "
        f"'{pipeline_outputs['output_workbook'].name}' → '{outputs2['output_workbook'].name}'"
    )


# ---------------------------------------------------------------------------
# Test — write_excel_workbook direct unit test (headers from in-memory data)
# ---------------------------------------------------------------------------

def test_write_excel_workbook_headers_from_in_memory_data(tmp_path: Path) -> None:
    """``write_excel_workbook`` with minimal in-memory data produces correct headers.

    This is a focused unit test that does not depend on the pipeline.
    """
    wb_path = tmp_path / "test_minimal.xlsx"

    route = RouteDecision(
        source_file="test.xlsx",
        product_id="PRODUCT_001",
        association_code="XXX001",
        custodian_name_chinese="测试",
        custodian_id="CUST001",
        custodian_name="test_custodian",
        adapter_key="test",
        route_source="mapping(product_id)",
        route_status="success",
        route_message="",
    )
    subject = SubjectRecord(
        source_file="test.xlsx",
        broker="测试券商",
        subject_code="1001",
        subject_name="测试科目",
    )
    position = PositionRecord(
        source_file="test.xlsx",
        broker="测试券商",
        instrument_name="测试标的",
        quantity=100.0,
        cost=50000.0,
        market_value=55000.0,
    )
    review_item = ReviewItem(
        source_file="test.xlsx",
        broker="测试券商",
        valuation_date="2025-03-27",
        raw_row_index=1,
        subject_code="1001",
        subject_name="测试科目",
        review_reason="测试复核原因",
    )

    write_excel_workbook(
        wb_path,
        routes=[route],
        subjects=[subject],
        positions=[position],
        review_items=[review_item],
    )

    sheets = _read_workbook_sheets(wb_path)

    # Verify sheet presence and order
    assert list(sheets.keys()) == EXPECTED_SHEET_NAMES

    # Verify headers match field constants
    for sheet_name, expected_fields in SHEET_FIELD_MAP.items():
        header_row = sheets[sheet_name][0]
        header_values = [str(v) for v in header_row]
        assert header_values == expected_fields, (
            f"{sheet_name} header mismatch in unit test"
        )

    # Verify data row counts (1 data row each)
    for sheet_name in EXPECTED_SHEET_NAMES:
        assert len(sheets[sheet_name]) == 2, (
            f"{sheet_name}: expected 1 header + 1 data row, got {len(sheets[sheet_name])} rows"
        )
