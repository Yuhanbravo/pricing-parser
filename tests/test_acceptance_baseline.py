from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from valuation_parser.pipeline import run_pipeline


EXPECTED_DIR = Path("data_samples/expected")


def test_strict_default_run_matches_acceptance_baseline(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"

    outputs = run_pipeline(Path("data_samples/raw"), Path("产品与托管机构映射表.csv"), output_dir)

    assert _normalized_text(outputs["routing_results"], encoding="utf-8-sig") == _normalized_text(EXPECTED_DIR / "routing_results.csv", encoding="utf-8-sig")
    assert _normalized_text(outputs["valuation_subjects"], encoding="utf-8-sig") == _normalized_text(EXPECTED_DIR / "valuation_subjects.csv", encoding="utf-8-sig")
    assert _normalized_text(outputs["valuation_positions"], encoding="utf-8-sig") == _normalized_text(EXPECTED_DIR / "valuation_positions.csv", encoding="utf-8-sig")
    assert _normalized_text(outputs["review_items"], encoding="utf-8-sig") == _normalized_text(EXPECTED_DIR / "review_items.csv", encoding="utf-8-sig")
    assert _normalized_text(outputs["parse_summary"], encoding="utf-8") == _normalized_text(EXPECTED_DIR / "parse_summary.md", encoding="utf-8")
    assert _read_workbook_rows(outputs["output_workbook"]) == _read_workbook_rows(EXPECTED_DIR / "估值表解析_output_2025-03-27.xlsx")


def _normalized_text(path: Path, *, encoding: str) -> str:
    return path.read_text(encoding=encoding).replace("\r\n", "\n")


def _read_workbook_rows(path: Path) -> dict[str, list[list[object | None]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            sheet_name: [list(row) for row in workbook[sheet_name].iter_rows(values_only=True)]
            for sheet_name in workbook.sheetnames
        }
    finally:
        workbook.close()