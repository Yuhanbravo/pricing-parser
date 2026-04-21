from __future__ import annotations

from abc import ABC
import csv
from pathlib import Path
import re

from openpyxl import load_workbook
import xlrd
import yaml

from valuation_parser.adapters.base import BaseValuationAdapter, annotate_subject_hierarchy, build_positions_and_review_items, enrich_subject_record
from valuation_parser.models import ParseArtifacts, RouteDecision, SubjectRecord


class SheetView(ABC):
    title: str

    def row_values(self, row_number: int) -> list[object]:
        raise NotImplementedError

    def iter_rows(self, start_row: int):
        raise NotImplementedError


class XlsxSheetView(SheetView):
    def __init__(self, worksheet) -> None:
        self._worksheet = worksheet
        self.title = worksheet.title

    def row_values(self, row_number: int) -> list[object]:
        return [cell for cell in next(self._worksheet.iter_rows(min_row=row_number, max_row=row_number, values_only=True), ())]

    def iter_rows(self, start_row: int):
        for row_index, row in enumerate(self._worksheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            yield row_index, list(row)


class XlsSheetView(SheetView):
    def __init__(self, worksheet) -> None:
        self._worksheet = worksheet
        self.title = worksheet.name

    def row_values(self, row_number: int) -> list[object]:
        row_index = max(row_number - 1, 0)
        return [self._worksheet.cell_value(row_index, col_index) for col_index in range(self._worksheet.ncols)]

    def iter_rows(self, start_row: int):
        for row_index in range(max(start_row - 1, 0), self._worksheet.nrows):
            row = [self._worksheet.cell_value(row_index, col_index) for col_index in range(self._worksheet.ncols)]
            yield row_index + 1, row


class CsvSheetView(SheetView):
    def __init__(self, path: Path, rows: list[list[str]]) -> None:
        self._path = path
        self._rows = rows
        self.title = path.stem

    def row_values(self, row_number: int) -> list[object]:
        row_index = max(row_number - 1, 0)
        if row_index >= len(self._rows):
            return []
        return self._rows[row_index]

    def iter_rows(self, start_row: int):
        for row_index, row in enumerate(self._rows[max(start_row - 1, 0):], start=max(start_row, 1)):
            yield row_index, row


class ConfigurableTabularAdapter(BaseValuationAdapter):
    key = "generic"

    def __init__(self, config_name: str, config_path: Path | None = None) -> None:
        default_path = Path(__file__).resolve().parents[3] / "config" / "adapters" / f"{config_name}.yaml"
        self.config_path = config_path or default_path
        self.config = load_adapter_config(self.config_path)

    def parse(self, source_file: Path, route: RouteDecision) -> ParseArtifacts:
        sheet = load_first_sheet(source_file)
        valuation_date = extract_valuation_date(
            sheet,
            scan_rows=int(self.config.get("valuation_date_scan_rows", 4)),
            preferred_row=int(self.config.get("valuation_date_row", 3)),
        )
        header_map = build_header_map(
            sheet,
            header_row=int(self.config.get("header_row", 4)),
            aliases=self.config.get("column_aliases", {}),
            fixed_columns=self.config.get("column_indices", {}),
        )
        subjects = extract_subjects(
            sheet=sheet,
            header_map=header_map,
            data_start_row=int(self.config.get("data_start_row", 5)),
            route=route,
            valuation_date=valuation_date,
            skip_name_keywords=self.config.get("skip_name_keywords", []),
        )
        subjects = [enrich_subject_record(subject) for subject in subjects]
        subjects = annotate_subject_hierarchy(subjects)
        subjects, positions, review_items = build_positions_and_review_items(subjects)
        return ParseArtifacts(route=route, subjects=subjects, positions=positions, review_items=review_items)


def load_adapter_config(config_path: Path) -> dict[str, object]:
    with config_path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle) or {}
    return data if isinstance(data, dict) else {}


def load_first_sheet(source_file: Path) -> SheetView:
    if source_file.suffix.lower() == ".xlsx":
        workbook = load_workbook(source_file, read_only=False, data_only=True)
        return XlsxSheetView(workbook[workbook.sheetnames[0]])

    if source_file.suffix.lower() == ".csv":
        return CsvSheetView(source_file, _read_csv_rows(source_file))

    workbook = xlrd.open_workbook(str(source_file))
    return XlsSheetView(workbook.sheet_by_index(0))


def _read_csv_rows(path: Path) -> list[list[str]]:
    encodings = ("utf-8-sig", "gb18030")
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.reader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"Unable to decode CSV workbook file: {path}") from last_error


def extract_valuation_date(sheet: SheetView, *, scan_rows: int, preferred_row: int) -> str | None:
    candidate_rows = [preferred_row]
    candidate_rows.extend(row_number for row_number in range(1, scan_rows + 1) if row_number != preferred_row)
    for row_number in candidate_rows:
        values = [cell_to_text(value) for value in sheet.row_values(row_number)]
        text = " ".join(value for value in values if value)
        if not text:
            continue
        hyphen_match = re.search(r"(\d{4}-\d{2}-\d{2})", text)
        if hyphen_match:
            return hyphen_match.group(1)
        compact_match = re.search(r"(\d{8})", text)
        if compact_match:
            digits = compact_match.group(1)
            return f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
    return None


def build_header_map(
    sheet: SheetView,
    *,
    header_row: int,
    aliases: dict[str, list[str]],
    fixed_columns: dict[str, int] | None = None,
) -> dict[str, int]:
    header_map: dict[str, int] = {}
    fixed_columns = fixed_columns or {}
    for field_name, column_index in fixed_columns.items():
        header_map[field_name] = int(column_index)

    headers = [normalize_header(value) for value in sheet.row_values(header_row)]
    for field_name, candidates in aliases.items():
        if field_name in header_map:
            continue
        normalized_candidates = {normalize_header(candidate) for candidate in candidates}
        for index, header in enumerate(headers):
            if header in normalized_candidates:
                header_map[field_name] = index
                break
    return header_map


def extract_subjects(
    *,
    sheet: SheetView,
    header_map: dict[str, int],
    data_start_row: int,
    route: RouteDecision,
    valuation_date: str | None,
    skip_name_keywords: list[str],
) -> list[SubjectRecord]:
    subjects: list[SubjectRecord] = []
    for row_index, row in sheet.iter_rows(data_start_row):
        values = [cell_to_text(value) for value in row]
        subject_code = get_value(values, header_map, "subject_code")
        subject_name = get_value(values, header_map, "subject_name")

        if not subject_code and not subject_name:
            continue
        if not is_subject_code(subject_code):
            continue
        if subject_name and any(keyword in subject_name for keyword in skip_name_keywords):
            continue

        subjects.append(
            SubjectRecord(
                source_file=str(route.source_file),
                sheet_name=sheet.title,
                valuation_date=valuation_date,
                product_id=route.product_id,
                association_code=route.association_code,
                custodian_id=route.custodian_id,
                custodian_name=route.custodian_name,
                adapter_key=route.adapter_key,
                route_source=route.route_source,
                subject_code=subject_code,
                subject_name=subject_name,
                quantity=parse_number(get_value(values, header_map, "quantity")),
                unit_cost=parse_number(get_value(values, header_map, "unit_cost")),
                cost=parse_number(get_value(values, header_map, "cost")),
                cost_pct_nav=parse_number(get_value(values, header_map, "cost_pct_nav")),
                market_price=parse_number(get_value(values, header_map, "market_price")),
                market_value=parse_number(get_value(values, header_map, "market_value")),
                market_value_pct_nav=parse_number(get_value(values, header_map, "market_value_pct_nav")),
                pnl=parse_number(get_value(values, header_map, "pnl")),
                suspension_info=get_value(values, header_map, "suspension_info"),
                raw_row_index=row_index,
                raw_text=" | ".join(value for value in values if value),
            )
        )
    return subjects


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def cell_to_text(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def get_value(values: list[str | None], header_map: dict[str, int], field_name: str) -> str | None:
    index = header_map.get(field_name)
    if index is None or index >= len(values):
        return None
    value = values[index]
    return value if value not in (None, "") else None


def parse_number(value: str | None) -> float | None:
    if value in (None, ""):
        return None
    normalized = value.replace(",", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def is_subject_code(value: str | None) -> bool:
    if not value:
        return False
    normalized = value.strip().upper()
    return bool(normalized and re.fullmatch(r"[0-9A-Z._\- ]+", normalized) and re.search(r"\d", normalized))