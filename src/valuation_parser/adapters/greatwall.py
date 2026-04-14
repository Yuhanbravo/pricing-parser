from __future__ import annotations

from pathlib import Path
import re

from openpyxl import load_workbook
import yaml

from valuation_parser.adapters.base import BaseValuationAdapter, annotate_subject_hierarchy, build_positions_and_review_items, enrich_subject_record
from valuation_parser.models import ParseArtifacts, RouteDecision, SubjectRecord


class GreatwallValuationAdapter(BaseValuationAdapter):
    key = "greatwall"

    def __init__(self, config_path: Path | None = None) -> None:
        default_path = Path(__file__).resolve().parents[3] / "config" / "adapters" / "greatwall.yaml"
        self.config_path = config_path or default_path
        self.config = _load_config(self.config_path)

    def parse(self, source_file: Path, route: RouteDecision) -> ParseArtifacts:
        workbook = load_workbook(source_file, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        header_row = int(self.config.get("header_row", 4))
        data_start_row = int(self.config.get("data_start_row", header_row + 1))
        header_map = _build_header_map(worksheet, header_row, self.config.get("column_aliases", {}))
        valuation_date = _extract_valuation_date(worksheet)

        subjects = _extract_subjects(
            worksheet=worksheet,
            header_map=header_map,
            data_start_row=data_start_row,
            route=route,
            valuation_date=valuation_date,
            skip_name_keywords=self.config.get("skip_name_keywords", []),
        )
        subjects = [enrich_subject_record(subject) for subject in subjects]
        subjects = annotate_subject_hierarchy(subjects)
        positions, review_items = build_positions_and_review_items(subjects)

        return ParseArtifacts(route=route, subjects=subjects, positions=positions, review_items=review_items)


def _load_config(config_path: Path) -> dict[str, object]:
    with config_path.open("r", encoding="utf-8") as handle:
        data = yaml.safe_load(handle) or {}
    return data if isinstance(data, dict) else {}


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip())


def _build_header_map(worksheet, header_row: int, aliases: dict[str, list[str]]) -> dict[str, int]:
    headers = [_normalize_header(cell) for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    header_map: dict[str, int] = {}
    for field_name, candidates in aliases.items():
        normalized_candidates = {_normalize_header(candidate) for candidate in candidates}
        for index, header in enumerate(headers):
            if header in normalized_candidates:
                header_map[field_name] = index
                break
    return header_map


def _extract_valuation_date(worksheet) -> str | None:
    row_values = next(worksheet.iter_rows(min_row=3, max_row=3, values_only=True))
    text = " ".join(str(value).strip() for value in row_values if value not in (None, ""))
    match = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if match:
        return match.group(1)
    return None


def _extract_subjects(
    *,
    worksheet,
    header_map: dict[str, int],
    data_start_row: int,
    route: RouteDecision,
    valuation_date: str | None,
    skip_name_keywords: list[str],
) -> list[SubjectRecord]:
    subjects: list[SubjectRecord] = []
    for row_index, row in enumerate(worksheet.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        values = [None if value is None else str(value).strip() for value in row]
        subject_code = _get_value(values, header_map, "subject_code")
        subject_name = _get_value(values, header_map, "subject_name")

        if not subject_code and not subject_name:
            continue
        if not _is_subject_code(subject_code):
            continue
        if subject_name and any(keyword in subject_name for keyword in skip_name_keywords):
            continue

        subject = SubjectRecord(
            source_file=str(route.source_file),
            sheet_name=worksheet.title,
            valuation_date=valuation_date,
            product_id=route.product_id,
            association_code=route.association_code,
            custodian_id=route.custodian_id,
            custodian_name=route.custodian_name,
            adapter_key=route.adapter_key,
            route_source=route.route_source,
            subject_code=subject_code,
            subject_name=subject_name,
            quantity=_parse_number(_get_value(values, header_map, "quantity")),
            unit_cost=_parse_number(_get_value(values, header_map, "unit_cost")),
            cost=_parse_number(_get_value(values, header_map, "cost")),
            cost_pct_nav=_parse_number(_get_value(values, header_map, "cost_pct_nav")),
            market_price=_parse_number(_get_value(values, header_map, "market_price")),
            market_value=_parse_number(_get_value(values, header_map, "market_value")),
            market_value_pct_nav=_parse_number(_get_value(values, header_map, "market_value_pct_nav")),
            pnl=_parse_number(_get_value(values, header_map, "pnl")),
            suspension_info=_get_value(values, header_map, "suspension_info"),
            raw_row_index=row_index,
            raw_text=" | ".join(value for value in values if value),
        )
        subjects.append(subject)

    return subjects


def _get_value(values: list[str | None], header_map: dict[str, int], field_name: str) -> str | None:
    if field_name not in header_map:
        return None
    index = header_map[field_name]
    if index >= len(values):
        return None
    value = values[index]
    return value if value not in (None, "") else None


def _parse_number(value: str | None) -> float | None:
    if value in (None, ""):
        return None
    normalized = value.replace(",", "")
    try:
        return float(normalized)
    except ValueError:
        return None


def _is_subject_code(value: str | None) -> bool:
    if not value:
        return False
    return bool(re.fullmatch(r"[0-9A-Z]+", value))
