from __future__ import annotations

from pathlib import Path
import csv
import re

from valuation_parser.models import ProductIdentity, WorkbookPreview

PRODUCT_PATTERN = re.compile(r"(?<![A-Z0-9])PRODUCT[_-]?(\d{3,})(?!\d)", re.IGNORECASE)
ASSOCIATION_PATTERN = re.compile(r"(?<![A-Z0-9])XXX[_-]?(\d{3,})(?!\d)", re.IGNORECASE)
CUSTODIAN_ALIAS_PATTERNS = {
    "国泰": "国泰",
}

def extract_product_identity(source_file: str | Path, preview: WorkbookPreview | None = None) -> ProductIdentity:
    path = Path(source_file)
    text_candidates = [("filename", path.stem)]
    top_row_texts: list[str] = []

    if preview is not None:
        text_candidates.extend(("sheet", value) for value in preview.sheet_names)
        text_candidates.extend(("header", value) for value in preview.header_texts)
        top_row_texts = preview.header_texts[:3]

    product_id = None
    association_code = None
    custodian_name_chinese = _extract_custodian_name_chinese(top_row_texts)
    evidence: list[str] = []

    for source, text in text_candidates:
        if product_id is None:
            product_match = PRODUCT_PATTERN.search(text)
            if product_match:
                product_id = f"PRODUCT_{product_match.group(1).zfill(3)}"
                evidence.append(f"product_id:{source}")

        if association_code is None:
            association_match = ASSOCIATION_PATTERN.search(text)
            if association_match:
                association_code = f"XXX{association_match.group(1).zfill(3)}"
                evidence.append(f"association_code:{source}")

        if product_id and association_code:
            break

    if custodian_name_chinese:
        evidence.append("custodian_name_chinese:header")

    if product_id and association_code:
        message = "resolved product_id and association_code"
    elif product_id:
        message = "resolved product_id only"
    elif association_code:
        message = "resolved association_code only"
    elif custodian_name_chinese:
        message = "resolved custodian_name_chinese only"
    else:
        message = "unable to resolve product identity from filename, sheet names, or header preview"

    return ProductIdentity(
        product_id=product_id,
        association_code=association_code,
        custodian_name_chinese=custodian_name_chinese,
        route_message=message,
        evidence=tuple(evidence),
    )


def _extract_custodian_name_chinese(top_row_texts: list[str]) -> str | None:
    for text in top_row_texts:
        for marker, canonical_name in CUSTODIAN_ALIAS_PATTERNS.items():
            if marker in text:
                return canonical_name
    return None


def preview_workbook(source_file: str | Path) -> WorkbookPreview:
    path = Path(source_file)
    suffix = path.suffix.lower()

    if suffix == ".csv":
        return _preview_csv(path)
    if suffix == ".xlsx":
        return _preview_xlsx(path)
    if suffix == ".xls":
        return _preview_xls(path)
    return WorkbookPreview(errors=[f"unsupported file type: {suffix}"])


def _preview_csv(path: Path) -> WorkbookPreview:
    header_texts: list[str] = []

    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                header_texts.append(" ".join(cell.strip() for cell in row if cell))
                if index >= 4:
                    break
    except UnicodeDecodeError:
        with path.open("r", encoding="gb18030", newline="") as handle:
            reader = csv.reader(handle)
            for index, row in enumerate(reader):
                header_texts.append(" ".join(cell.strip() for cell in row if cell))
                if index >= 4:
                    break
    except OSError as exc:
        return WorkbookPreview(errors=[str(exc)])

    return WorkbookPreview(sheet_names=[path.stem], header_texts=header_texts)


def _preview_xlsx(path: Path) -> WorkbookPreview:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return WorkbookPreview(errors=[f"openpyxl is required to inspect .xlsx files: {exc}"])

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - defensive path for malformed workbooks
        return WorkbookPreview(errors=[str(exc)])

    sheet_names = list(workbook.sheetnames)
    header_texts: list[str] = []

    for sheet_name in sheet_names[:3]:
        worksheet = workbook[sheet_name]
        for row in worksheet.iter_rows(min_row=1, max_row=3, max_col=8, values_only=True):
            text = " ".join(str(value).strip() for value in row if value not in (None, ""))
            if text:
                header_texts.append(text)

    return WorkbookPreview(sheet_names=sheet_names, header_texts=header_texts)


def _preview_xls(path: Path) -> WorkbookPreview:
    try:
        import xlrd
    except ImportError as exc:
        return WorkbookPreview(errors=[f"xlrd is required to inspect .xls files: {exc}"])

    try:
        workbook = xlrd.open_workbook(path)
    except Exception as exc:  # pragma: no cover - defensive path for malformed workbooks
        return WorkbookPreview(errors=[str(exc)])

    sheet_names = workbook.sheet_names()
    header_texts: list[str] = []

    for sheet_name in sheet_names[:3]:
        sheet = workbook.sheet_by_name(sheet_name)
        max_rows = min(sheet.nrows, 3)
        max_cols = min(sheet.ncols, 8)
        for row_index in range(max_rows):
            values = [str(sheet.cell_value(row_index, col_index)).strip() for col_index in range(max_cols)]
            text = " ".join(value for value in values if value)
            if text:
                header_texts.append(text)

    return WorkbookPreview(sheet_names=sheet_names, header_texts=header_texts)
