from __future__ import annotations

from collections.abc import Iterable
import csv
from itertools import zip_longest
from pathlib import Path
import re

from valuation_parser.adapter_registry import SUPPORTED_ADAPTERS
from valuation_parser.models import MappingRecord

CANONICAL_FIELDS = {
    "product_id",
    "association_code",
    "custodian_id",
    "custodian_name",
    "adapter_key",
}

COMPACT_FIELDS = {
    "fake_custodian_id",
    "fake_custodian_name",
    "true_custodian_name",
    "fake_product_ids",
    "fake_association_codes",
}

NAME_TO_ADAPTER_KEY = {
    "兴业证券": "xyzc",
    "招商证券": "cmsc",
    "中信证券": "citics",
    "东方证券": "orient",
    "中信建投": "csc",
    "国泰海通": "gtja",
    "国信证券": "guosen",
    "长城证券": "greatwall",
}


def load_mapping(mapping_file: str | Path, include_inactive: bool = False) -> list[MappingRecord]:
    rows = _read_rows(Path(mapping_file))
    if not rows:
        return []

    normalized_headers = {header.strip() for header in rows[0].keys()}
    if CANONICAL_FIELDS.issubset(normalized_headers):
        records = _load_canonical_rows(rows)
    elif COMPACT_FIELDS.issubset(normalized_headers):
        records = _load_compact_rows(rows)
    else:
        raise ValueError("Unsupported mapping schema. Expected canonical fields or compact sample fields.")

    active_records = [record for record in records if record.is_active or include_inactive]
    _validate_records(active_records)
    return active_records


def _read_rows(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_rows(path)
    if suffix == ".xlsx":
        return _read_xlsx_rows(path)
    raise ValueError(f"Unsupported mapping file type: {path.suffix}")


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    encodings = ("utf-8-sig", "gb18030")
    last_error: Exception | None = None

    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError as exc:
            last_error = exc

    raise ValueError(f"Unable to decode CSV mapping file: {path}") from last_error


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("openpyxl is required to read xlsx mapping files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, str]] = []
    for values in rows[1:]:
        if not any(value not in (None, "") for value in values):
            continue
        record = {
            headers[index]: "" if value is None else str(value).strip()
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        records.append(record)

    return records


def _load_canonical_rows(rows: Iterable[dict[str, str]]) -> list[MappingRecord]:
    records: list[MappingRecord] = []
    for row in rows:
        records.append(
            MappingRecord(
                product_id=_normalize_product_id(row.get("product_id")),
                association_code=_normalize_association_code(row.get("association_code")),
                custodian_id=(row.get("custodian_id") or "").strip(),
                custodian_name=(row.get("custodian_name") or "").strip(),
                true_custodian_name=(row.get("true_custodian_name") or row.get("custodian_name") or "").strip() or None,
                adapter_key=(row.get("adapter_key") or "generic").strip().lower(),
                is_active=_parse_bool(row.get("is_active"), default=True),
                note=(row.get("note") or "").strip() or None,
            )
        )
    return records


def _load_compact_rows(rows: Iterable[dict[str, str]]) -> list[MappingRecord]:
    records: list[MappingRecord] = []
    for row in rows:
        product_ids = _split_values(row.get("fake_product_ids"))
        association_codes = _split_values(row.get("fake_association_codes"))
        custodian_id = (row.get("fake_custodian_id") or "").strip()
        custodian_name = (row.get("fake_custodian_name") or "").strip()
        true_name = (row.get("true_custodian_name") or "").strip()
        adapter_key = _derive_adapter_key(true_name or custodian_name)

        for product_id, association_code in zip_longest(product_ids, association_codes, fillvalue=""):
            if not product_id and not association_code:
                continue
            records.append(
                MappingRecord(
                    product_id=_normalize_product_id(product_id),
                    association_code=_normalize_association_code(association_code),
                    custodian_id=custodian_id,
                    custodian_name=custodian_name,
                    true_custodian_name=true_name or None,
                    adapter_key=adapter_key,
                    is_active=True,
                    note=true_name or None,
                )
            )
    return records


def _split_values(value: str | None) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in value.split(";") if item.strip()]


def _normalize_product_id(value: str | None) -> str | None:
    if not value:
        return None
    match = re.search(r"PRODUCT[_-]?(\d{3,})", value, re.IGNORECASE)
    if match:
        return f"PRODUCT_{match.group(1).zfill(3)}"
    return value.strip().upper()


def _normalize_association_code(value: str | None) -> str | None:
    if not value:
        return None
    match = re.search(r"XXX[_-]?(\d{3,})", value, re.IGNORECASE)
    if match:
        return f"XXX{match.group(1).zfill(3)}"
    return value.strip().upper()


def _derive_adapter_key(name: str) -> str:
    for marker, adapter_key in NAME_TO_ADAPTER_KEY.items():
        if marker in name:
            return adapter_key
    return "generic"


def _parse_bool(value: str | None, default: bool) -> bool:
    if value is None or value == "":
        return default
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y"}:
        return True
    if normalized in {"0", "false", "no", "n"}:
        return False
    return default


def _validate_records(records: Iterable[MappingRecord]) -> None:
    pair_index: set[tuple[str | None, str | None]] = set()
    product_index: set[str] = set()
    association_index: set[str] = set()

    for record in records:
        if record.adapter_key.strip() == "":
            raise ValueError(f"Missing adapter_key for custodian_id={record.custodian_id}")
        if record.adapter_key not in SUPPORTED_ADAPTERS:
            raise ValueError(f"Unsupported adapter_key in mapping: {record.adapter_key}")

        pair_key = (record.product_id, record.association_code)
        if pair_key in pair_index:
            raise ValueError(f"Duplicate mapping pair: {pair_key}")
        pair_index.add(pair_key)

        if record.product_id:
            if record.product_id in product_index:
                raise ValueError(f"Duplicate product_id mapping: {record.product_id}")
            product_index.add(record.product_id)

        if record.association_code:
            if record.association_code in association_index:
                raise ValueError(f"Duplicate association_code mapping: {record.association_code}")
            association_index.add(record.association_code)
